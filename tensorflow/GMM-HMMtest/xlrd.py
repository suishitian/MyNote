from openpyxl import load_workbook
import test2

wb = load_workbook("/home/shitian/Documents/MyGithub/ASRtest/test1.xlsx")
print(wb.sheetnames)
path = "./video_7m/"
cpath = "audioNameForCPP.txt"
sheet = wb[wb.sheetnames[0]]
print(sheet.title)
print(sheet["A1"].value)
output = open('thefile.txt', 'w')
coutput = open('caudio.txt', 'w')
for row in sheet.rows:
    if row[0].value=="audio":
        continue
    filename = path+row[0].value
    coutput.writelines(row[0].value+"\n")
    str = row[0].value+"/"+test2.asr(filename,"zh-CHS",1,16000,"wav")
    output.writelines(str+"\n")
    print str
output.close()
coutput.close()
